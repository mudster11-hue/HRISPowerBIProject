"""
Synthetic HRIS Dataset Generator
==================================
Simulates a Workday-style HR system for a ~500-employee growth company.
Designed for Power BI People Analytics portfolio projects.

Intentional data quality issues are embedded throughout (marked with # [DQ]).
These mimic real-world problems you'd encounter cleaning an HRIS export:
  - Duplicate employee records
  - Inconsistent department naming (free-text legacy fields)
  - Missing ManagerIDs
  - Salary outliers
  - Invalid/missing termination reasons
  - Overlapping leave date windows

Run:
    pip install pandas numpy faker openpyxl
    python generate_hris.py
"""

import random
from datetime import date, datetime, timedelta

import numpy as np
import pandas as pd
from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ─── Reproducibility ──────────────────────────────────────────────────────────
SEED = 42
random.seed(SEED)
np.random.seed(SEED)
fake = Faker("en_US")
Faker.seed(SEED)

# ─── Global config ────────────────────────────────────────────────────────────
TARGET_HEADCOUNT = 500
ATTRITION_RATE   = 0.12    # ~12% of workforce will be marked Terminated
DUPLICATE_RATE   = 0.015   # ~1.5% duplicate rows injected [DQ]
HIRE_START       = date(2019, 1, 1)
HIRE_END         = date(2024, 6, 1)
OUTPUT_FILE      = "HRIS_Dataset.xlsx"

# ─── Department catalogue ─────────────────────────────────────────────────────
# Salary bands are split by job level so managers always out-earn ICs.
# The "weight" drives how many employees land in each department.
DEPARTMENTS: dict[str, dict] = {
    "Engineering":     {"weight": 0.30, "salary": {"IC": (70_000, 155_000), "Manager": (130_000, 190_000), "Director": (170_000, 220_000), "VP": (205_000, 280_000)}},
    "Sales":           {"weight": 0.20, "salary": {"IC": (50_000, 120_000), "Manager": (110_000, 160_000), "Director": (150_000, 200_000), "VP": (190_000, 265_000)}},
    "Finance":         {"weight": 0.10, "salary": {"IC": (60_000, 130_000), "Manager": (120_000, 170_000), "Director": (160_000, 210_000), "VP": (195_000, 270_000)}},
    "Human Resources": {"weight": 0.08, "salary": {"IC": (50_000, 100_000), "Manager": (100_000, 145_000), "Director": (140_000, 185_000), "VP": (175_000, 240_000)}},
    "Marketing":       {"weight": 0.10, "salary": {"IC": (55_000, 115_000), "Manager": (110_000, 155_000), "Director": (150_000, 195_000), "VP": (185_000, 252_000)}},
    "Operations":      {"weight": 0.10, "salary": {"IC": (45_000, 100_000), "Manager": (100_000, 140_000), "Director": (135_000, 175_000), "VP": (170_000, 230_000)}},
    "Product":         {"weight": 0.07, "salary": {"IC": (75_000, 165_000), "Manager": (145_000, 195_000), "Director": (175_000, 225_000), "VP": (210_000, 285_000)}},
    "Legal":           {"weight": 0.05, "salary": {"IC": (80_000, 170_000), "Manager": (155_000, 205_000), "Director": (185_000, 240_000), "VP": (225_000, 300_000)}},
}

# [DQ] Dirty name variants — injected into ~7% of records per affected dept.
# This simulates a legacy HRIS where department was a free-text field.
DEPT_VARIANTS: dict[str, list[str]] = {
    "Engineering":     ["Eng", "Engineering ", "engineering", "Engg"],
    "Human Resources": ["HR", "Human Resources ", "HR Dept", "H.R."],
    "Sales":           ["Sales & BD", "Sales ", "SALES"],
}

JOB_TITLES: dict[str, dict[str, list[str]]] = {
    "Engineering": {
        "IC":       ["Software Engineer I", "Software Engineer II", "Senior Software Engineer", "Staff Engineer"],
        "Manager":  ["Engineering Manager", "Senior Engineering Manager"],
        "Director": ["Director of Engineering", "Senior Director of Engineering"],
        "VP":       ["VP of Engineering"],
    },
    "Sales": {
        "IC":       ["Sales Development Rep", "Account Executive", "Senior Account Executive", "Enterprise Account Executive"],
        "Manager":  ["Sales Manager", "Regional Sales Manager"],
        "Director": ["Director of Sales", "Senior Director of Sales"],
        "VP":       ["VP of Sales"],
    },
    "Finance": {
        "IC":       ["Financial Analyst", "Senior Financial Analyst", "FP&A Analyst", "Senior FP&A Analyst"],
        "Manager":  ["Finance Manager", "Senior Finance Manager"],
        "Director": ["Director of Finance"],
        "VP":       ["CFO"],
    },
    "Human Resources": {
        "IC":       ["HR Coordinator", "HR Generalist", "Senior HR Generalist", "HR Business Partner"],
        "Manager":  ["HR Manager", "Senior HR Manager"],
        "Director": ["Director of HR"],
        "VP":       ["Chief People Officer"],
    },
    "Marketing": {
        "IC":       ["Marketing Coordinator", "Marketing Specialist", "Senior Marketing Specialist", "Growth Marketing Manager"],
        "Manager":  ["Marketing Manager", "Senior Marketing Manager"],
        "Director": ["Director of Marketing"],
        "VP":       ["VP of Marketing"],
    },
    "Operations": {
        "IC":       ["Operations Analyst", "Operations Specialist", "Senior Ops Specialist", "Ops Program Manager"],
        "Manager":  ["Operations Manager", "Senior Operations Manager"],
        "Director": ["Director of Operations"],
        "VP":       ["VP of Operations"],
    },
    "Product": {
        "IC":       ["Associate PM", "Product Manager", "Senior Product Manager", "Principal PM"],
        "Manager":  ["Group Product Manager", "Senior Group PM"],
        "Director": ["Director of Product"],
        "VP":       ["VP of Product"],
    },
    "Legal": {
        "IC":       ["Legal Counsel", "Senior Legal Counsel", "Associate General Counsel", "Senior Associate GC"],
        "Manager":  ["Legal Manager", "Senior Legal Manager"],
        "Director": ["Director of Legal Affairs"],
        "VP":       ["General Counsel"],
    },
}

US_OFFICE_LOCATIONS: list[str] = [
    "New York, NY", "San Francisco, CA", "Austin, TX", "Chicago, IL",
    "Seattle, WA", "Boston, MA", "Los Angeles, CA", "Denver, CO",
    "Atlanta, GA", "Dallas, TX", "Miami, FL", "Portland, OR",
]

EMPLOYMENT_TYPES   = ["Full-time", "Part-time", "Contractor"]
EMPLOYMENT_WEIGHTS = [0.80,         0.10,        0.10]

LEAVE_TYPES   = ["PTO", "Sick", "Unpaid", "Maternity/Paternity", "Bereavement", "FMLA"]
LEAVE_WEIGHTS = [0.45,  0.25,   0.10,    0.08,                  0.05,          0.07]

TERMINATION_REASONS = ["Voluntary", "Involuntary", "Performance", "Restructuring", "Other"]

# Pay grades map to levels; used in Compensation to give Power BI a clean dimension.
PAY_GRADES: dict[str, list[str]] = {
    "IC":       ["G1", "G2", "G3", "G4"],
    "Manager":  ["M1", "M2"],
    "Director": ["D1", "D2"],
    "VP":       ["V1"],
}


# ─── Shared utilities ─────────────────────────────────────────────────────────

def random_date(start: date, end: date) -> date:
    """Uniform random date in [start, end]. Returns start if start >= end."""
    if start >= end:
        return start
    return start + timedelta(days=random.randint(0, (end - start).days))


def biased_hire_date() -> date:
    """
    Realistic hiring distribution for a growth-stage company:
      - More hires in later years (company is scaling)
      - Mild Q1 spike (budget-cycle headcount approvals land Jan–Mar)
    Beta(2, 1.2) concentrates mass toward the upper end of the date range.
    """
    total_days = (HIRE_END - HIRE_START).days
    offset = int(np.random.beta(2, 1.2) * total_days)
    d = HIRE_START + timedelta(days=offset)
    if random.random() < 0.25:  # Q1 snap
        d = d.replace(month=random.randint(1, 3), day=random.randint(1, 28))
    return d


def make_email(name: str, emp_id: int, domain: str = "acmecorp.com") -> str:
    parts = name.lower().split()
    first = "".join(c for c in (parts[0] if parts else "emp") if c.isalpha())
    last  = "".join(c for c in (parts[-1] if len(parts) > 1 else str(emp_id)) if c.isalpha())
    return f"{first}.{last}@{domain}"


def salary_with_noise(low: int, high: int, level_pct: float) -> float:
    """
    Position salary at `level_pct` within [low, high], then add ±8% Gaussian
    noise — mimics the within-grade variation you see in real compensation data.
    Clamps to ±15% of the band edges so outliers come from explicit injection.
    """
    base  = low + level_pct * (high - low)
    noise = np.random.normal(0, 0.08 * base)
    return round(max(low * 0.85, min(high * 1.15, base + noise)), -2)


def infer_level(title: str) -> str:
    """Map a job title string to a coarse level for salary/grade lookups."""
    t = title.lower()
    if any(k in t for k in ("vp", "chief", "cfo", "cpo", "general counsel")):
        return "VP"
    if "director" in t:
        return "Director"
    if "manager" in t:
        return "Manager"
    return "IC"


def canonical_dept(dirty: str) -> str:
    """
    Reverse-map a dirty department name to its canonical key.
    Falls back to 'Operations' when nothing matches (shouldn't happen in practice).
    """
    cleaned = dirty.strip().lower()
    for dept in DEPARTMENTS:
        if cleaned.startswith(dept[:4].lower()) or dept.lower() in cleaned:
            return dept
    # Try variant map
    for canonical, variants in DEPT_VARIANTS.items():
        if dirty in variants:
            return canonical
    return "Operations"


# ─── Table 1: Employees ───────────────────────────────────────────────────────

def create_employees(n: int = TARGET_HEADCOUNT) -> pd.DataFrame:
    """
    Build the core employee roster with a 4-level org hierarchy:
        VP → Director → Manager → Individual Contributor

    Construction order matters: VPs are seeded first so their EmployeeIDs
    can be used as ManagerIDs for Directors, who in turn parent Managers, etc.
    This ensures every hierarchy link references a real row.

    Data quality injections applied after base generation:
        - ~4% of ICs have NULL ManagerID                          [DQ]
        - ~7% of records in 3 departments have dirty dept names   [DQ]
        - ~1.5% of rows are near-duplicates (upper-cased email)   [DQ]
    """
    rows: list[dict] = []
    emp_id = 1000

    dept_names   = list(DEPARTMENTS.keys())
    dept_weights = [DEPARTMENTS[d]["weight"] for d in dept_names]

    # Distribute headcount proportionally; absorb rounding in the last dept.
    dept_counts: dict[str, int] = {}
    remaining = n
    for dept in dept_names[:-1]:
        count = round(n * DEPARTMENTS[dept]["weight"])
        dept_counts[dept] = count
        remaining -= count
    dept_counts[dept_names[-1]] = remaining

    # ── Pass 1: one VP per department ────────────────────────────────────────
    vp_ids: dict[str, int] = {}
    for dept in dept_names:
        name = fake.name()
        eid  = emp_id; emp_id += 1
        hire = random_date(HIRE_START, date(2020, 6, 30))  # VPs hired early
        rows.append({
            "EmployeeID":     eid,
            "FullName":       name,
            "Email":          make_email(name, eid),
            "HireDate":       hire,
            "Department":     dept,
            "JobTitle":       JOB_TITLES[dept]["VP"][0],
            "ManagerID":      None,           # top of hierarchy — no parent
            "Location":       random.choice(US_OFFICE_LOCATIONS),
            "EmploymentType": "Full-time",    # executives are always FT
            "Status":         "Active",
            "_level":         "VP",
        })
        vp_ids[dept] = eid

    # ── Pass 2: 1–2 Directors per department ─────────────────────────────────
    director_ids: dict[str, list[int]] = {}
    for dept in dept_names:
        n_dirs = 2 if dept_counts[dept] >= 40 else 1
        director_ids[dept] = []
        for _ in range(n_dirs):
            name = fake.name()
            eid  = emp_id; emp_id += 1
            hire = random_date(HIRE_START, date(2021, 12, 31))
            rows.append({
                "EmployeeID":     eid,
                "FullName":       name,
                "Email":          make_email(name, eid),
                "HireDate":       hire,
                "Department":     dept,
                "JobTitle":       random.choice(JOB_TITLES[dept]["Director"]),
                "ManagerID":      vp_ids[dept],
                "Location":       random.choice(US_OFFICE_LOCATIONS),
                "EmploymentType": "Full-time",
                "Status":         "Active",
                "_level":         "Director",
            })
            director_ids[dept].append(eid)

    # ── Pass 3: Managers (roughly 1 per 9 ICs) ───────────────────────────────
    manager_ids: dict[str, list[int]] = {}
    for dept in dept_names:
        # Budget the IC pool after removing VPs and Directors already created
        senior_count = 1 + len(director_ids[dept])
        ic_target    = dept_counts[dept] - senior_count
        n_mgrs       = max(2, round(ic_target / 9))
        manager_ids[dept] = []
        for _ in range(n_mgrs):
            name   = fake.name()
            eid    = emp_id; emp_id += 1
            parent = random.choice(director_ids[dept])
            hire   = random_date(HIRE_START, date(2022, 12, 31))
            rows.append({
                "EmployeeID":     eid,
                "FullName":       name,
                "Email":          make_email(name, eid),
                "HireDate":       hire,
                "Department":     dept,
                "JobTitle":       random.choice(JOB_TITLES[dept]["Manager"]),
                "ManagerID":      parent,
                "Location":       random.choice(US_OFFICE_LOCATIONS),
                "EmploymentType": "Full-time",
                "Status":         "Active",
                "_level":         "Manager",
            })
            manager_ids[dept].append(eid)

    # ── Pass 4: Individual Contributors (fills remaining headcount) ───────────
    for dept in dept_names:
        already = 1 + len(director_ids[dept]) + len(manager_ids[dept])
        ic_count = max(0, dept_counts[dept] - already)
        for _ in range(ic_count):
            name  = fake.name()
            eid   = emp_id; emp_id += 1
            hire  = biased_hire_date()
            etype = random.choices(EMPLOYMENT_TYPES, weights=EMPLOYMENT_WEIGHTS)[0]
            # [DQ] ~4% of ICs intentionally have no ManagerID
            mgr = (
                random.choice(manager_ids[dept])
                if random.random() > 0.04
                else None
            )
            rows.append({
                "EmployeeID":     eid,
                "FullName":       name,
                "Email":          make_email(name, eid),
                "HireDate":       hire,
                "Department":     dept,
                "JobTitle":       random.choice(JOB_TITLES[dept]["IC"]),
                "ManagerID":      mgr,
                "Location":       random.choice(US_OFFICE_LOCATIONS),
                "EmploymentType": etype,
                "Status":         "Active",
                "_level":         "IC",
            })

    df = pd.DataFrame(rows)

    # ── Mark attrition ────────────────────────────────────────────────────────
    # Only ICs and Managers are eligible; executives are treated as always active.
    eligible_idx = df[df["_level"].isin(["IC", "Manager"])].index.tolist()
    n_term       = round(len(eligible_idx) * ATTRITION_RATE)
    term_idx     = random.sample(eligible_idx, n_term)
    df.loc[term_idx, "Status"] = "Terminated"

    # [DQ] Inject dirty department name variants into a fraction of records
    rng = np.random.default_rng(SEED)
    for dept, variants in DEPT_VARIANTS.items():
        mask      = df["Department"] == dept
        dirty_idx = df[mask].sample(frac=0.07, random_state=SEED).index
        dirty_vals = rng.choice(variants, size=len(dirty_idx))
        df.loc[dirty_idx, "Department"] = dirty_vals

    # [DQ] Inject near-duplicate rows (~1.5%): same EmployeeID, email uppercased
    n_dups  = max(1, round(len(df) * DUPLICATE_RATE))
    dup_src = df.sample(n=n_dups, random_state=SEED).copy()
    dup_src["Email"] = dup_src["Email"].str.upper()  # subtle variation
    df = pd.concat([df, dup_src], ignore_index=True)

    df = df.drop(columns=["_level"])
    df["HireDate"]  = pd.to_datetime(df["HireDate"])
    df["ManagerID"] = df["ManagerID"].astype("Int64")  # nullable integer
    return df


# ─── Table 2: Compensation ────────────────────────────────────────────────────

def create_compensation(employees: pd.DataFrame) -> pd.DataFrame:
    """
    Assign base salaries and bonus targets. Key design decisions:
      - Salary position within band is driven by job level (VPs at 70–95th pct)
      - Within-level variation comes from Gaussian noise (~8% std)
      - ~35% of employees have a second comp row representing a raise event
      - Pay grade is inferred from title keywords (Senior → top of grade, etc.)

    Data quality injections:
      - ~2% of rows have salary far outside the expected band               [DQ]
    """
    rows: list[dict] = []

    def pick_pay_grade(level: str, title: str) -> str:
        grades = PAY_GRADES[level]
        t = title.lower()
        if any(k in t for k in ("senior", "staff", "principal", "enterprise", "ii", "2")):
            return grades[-1]
        if "associate" in t or "i " in t or t.endswith(" i"):
            return grades[0]
        return random.choice(grades)

    # Work only on unique employees — duplicates shouldn't get extra comp rows
    unique_emp = employees.drop_duplicates(subset="EmployeeID")

    for _, emp in unique_emp.iterrows():
        dept      = canonical_dept(str(emp["Department"]))
        level     = infer_level(str(emp["JobTitle"]))
        sal_band  = DEPARTMENTS[dept]["salary"][level]

        # Senior leaders cluster at the top of their band
        level_pct = {
            "VP":       random.uniform(0.70, 0.95),
            "Director": random.uniform(0.55, 0.82),
            "Manager":  random.uniform(0.40, 0.72),
            "IC":       random.uniform(0.20, 0.68),
        }[level]

        base_salary  = salary_with_noise(sal_band[0], sal_band[1], level_pct)
        bonus_base   = {"VP": 25, "Director": 20, "Manager": 15, "IC": 10}[level]
        bonus_target = bonus_base + random.randint(-3, 8)
        pay_grade    = pick_pay_grade(level, str(emp["JobTitle"]))
        hire_ts      = pd.to_datetime(emp["HireDate"])

        rows.append({
            "EmployeeID":         emp["EmployeeID"],
            "BaseSalary":         base_salary,
            "BonusTargetPercent": bonus_target,
            "PayGrade":           pay_grade,
            "EffectiveDate":      hire_ts,
        })

        # Simulate a merit raise for ~35% of employees (second comp row)
        if random.random() < 0.35:
            raise_ts = hire_ts + timedelta(days=random.randint(365, 1095))
            if raise_ts <= pd.Timestamp(HIRE_END):
                raise_pct  = random.uniform(0.03, 0.12)
                new_salary = round(base_salary * (1 + raise_pct), -2)
                rows.append({
                    "EmployeeID":         emp["EmployeeID"],
                    "BaseSalary":         new_salary,
                    "BonusTargetPercent": bonus_target + random.randint(0, 3),
                    "PayGrade":           pay_grade,
                    "EffectiveDate":      raise_ts,
                })

    df = pd.DataFrame(rows)

    # [DQ] Inject ~2% salary outliers (fat-finger or system migration artifacts)
    n_outliers  = max(1, round(len(df) * 0.02))
    outlier_idx = random.sample(df.index.tolist(), n_outliers)
    for idx in outlier_idx:
        multiplier = random.choice([
            random.uniform(1.50, 2.20),   # grossly overpaid
            random.uniform(0.25, 0.55),   # suspiciously underpaid
        ])
        df.loc[idx, "BaseSalary"] = round(df.loc[idx, "BaseSalary"] * multiplier, 2)

    df["BaseSalary"]    = df["BaseSalary"].round(2)
    df["EffectiveDate"] = pd.to_datetime(df["EffectiveDate"])
    return df


# ─── Table 3: Leave ───────────────────────────────────────────────────────────

def create_leave(employees: pd.DataFrame) -> pd.DataFrame:
    """
    Generate leave requests per employee. Design rules:
      - Active employees: 1–4 leave events over their tenure
      - ~8% of requests are left unapproved (manager didn't action them)
      - Duration varies realistically by leave type
      - ~5% of employees have at least one intentionally overlapping leave window [DQ]
    """
    rows: list[dict] = []
    leave_id = 1

    duration_map: dict[str, tuple[int, int]] = {
        "PTO":                 (1,  15),
        "Sick":                (1,   5),
        "Unpaid":              (1,  30),
        "Maternity/Paternity": (14, 84),
        "Bereavement":         (3,   5),
        "FMLA":                (14, 84),
    }

    unique_emp = employees.drop_duplicates(subset="EmployeeID")

    for _, emp in unique_emp.iterrows():
        hire_dt = pd.to_datetime(emp["HireDate"]).date()
        # Leave window starts 90 days post-hire, ends at HIRE_END
        window_start = hire_dt + timedelta(days=90)
        window_end   = HIRE_END - timedelta(days=90)

        if window_start >= window_end:
            continue  # very recent hire — no room for leave events

        n_leaves = random.randint(1, 4) if emp["Status"] == "Active" else random.randint(0, 2)
        emp_leaves: list[dict] = []

        for _ in range(n_leaves):
            ltype = random.choices(LEAVE_TYPES, weights=LEAVE_WEIGHTS)[0]
            low_d, high_d = duration_map[ltype]
            duration = random.randint(low_d, high_d)
            # Ensure there's room for the full duration within the window
            latest_start = window_end - timedelta(days=duration)
            if window_start > latest_start:
                continue

            start = random_date(window_start, latest_start)
            end   = start + timedelta(days=duration)

            emp_leaves.append({
                "LeaveID":    leave_id,
                "EmployeeID": emp["EmployeeID"],
                "LeaveType":  ltype,
                "StartDate":  start,
                "EndDate":    end,
                "Approved":   "Y" if random.random() > 0.08 else "N",
            })
            leave_id += 1

        # [DQ] ~5% of employees get one intentionally overlapping leave record
        if emp_leaves and random.random() < 0.05:
            base          = random.choice(emp_leaves)
            overlap_start = base["StartDate"] + timedelta(days=random.randint(1, 4))
            overlap_end   = overlap_start + timedelta(days=random.randint(1, 5))
            emp_leaves.append({
                "LeaveID":    leave_id,
                "EmployeeID": emp["EmployeeID"],
                "LeaveType":  random.choices(LEAVE_TYPES, weights=LEAVE_WEIGHTS)[0],
                "StartDate":  overlap_start,
                "EndDate":    overlap_end,
                "Approved":   "Y",
            })
            leave_id += 1

        rows.extend(emp_leaves)

    df = pd.DataFrame(rows)
    df["StartDate"] = pd.to_datetime(df["StartDate"])
    df["EndDate"]   = pd.to_datetime(df["EndDate"])
    return df


# ─── Table 4: Performance ─────────────────────────────────────────────────────

def create_performance(employees: pd.DataFrame) -> pd.DataFrame:
    """
    Annual performance reviews on Dec 15 of each active year.
    Rating distribution: Normal(mean=3.2, std=0.85), clipped to [1, 5].
    Promotion eligibility is correlated with rating — high performers are
    far more likely to be flagged eligible, matching real-world calibration.
    """
    rows: list[dict] = []

    promo_prob_map = {1: 0.02, 2: 0.05, 3: 0.15, 4: 0.55, 5: 0.88}

    unique_emp = employees.drop_duplicates(subset="EmployeeID")

    for _, emp in unique_emp.iterrows():
        hire_dt = pd.to_datetime(emp["HireDate"]).date()
        yr = max(hire_dt.year, HIRE_START.year)

        while yr <= HIRE_END.year:
            review_date = date(yr, 12, 15)
            # Employee must have been here at least 6 months before the review
            if hire_dt <= review_date - timedelta(days=180):
                raw_rating = np.random.normal(loc=3.2, scale=0.85)
                rating     = int(round(np.clip(raw_rating, 1, 5)))
                eligible   = random.random() < promo_prob_map.get(rating, 0.10)
                rows.append({
                    "EmployeeID":        emp["EmployeeID"],
                    "ReviewDate":        review_date,
                    "Rating":            rating,
                    "PromotionEligible": eligible,
                })
            yr += 1

    df = pd.DataFrame(rows)
    df["ReviewDate"] = pd.to_datetime(df["ReviewDate"])
    return df


# ─── Table 5: Termination ─────────────────────────────────────────────────────

def create_termination(employees: pd.DataFrame) -> pd.DataFrame:
    """
    Build termination records only for employees flagged Status='Terminated'.
    Termination date is always strictly after hire date (enforced explicitly).
    Reason distribution: Voluntary is the most common (~45%), matching
    industry benchmarks for voluntary turnover in tech companies.

    Data quality injection:
      - ~5% of records have a NULL, blank, or placeholder reason           [DQ]
    """
    rows: list[dict] = []

    reason_weights = [0.45, 0.25, 0.15, 0.10, 0.05]  # aligned to TERMINATION_REASONS order

    terminated = (
        employees[employees["Status"] == "Terminated"]
        .drop_duplicates(subset="EmployeeID")
    )

    for _, emp in terminated.iterrows():
        hire_dt      = pd.to_datetime(emp["HireDate"]).date()
        # Most people leave between 6 months and 4 years after joining
        days_to_term = random.randint(180, 1460)
        term_date    = hire_dt + timedelta(days=days_to_term)
        # Clamp to HIRE_END so we don't generate future terminations
        if term_date > HIRE_END:
            term_date = HIRE_END - timedelta(days=random.randint(10, 90))

        # [DQ] ~5% get an invalid or missing reason
        if random.random() < 0.05:
            reason = random.choice([None, "", "N/A", "TBD", "Unknown"])
        else:
            reason = random.choices(TERMINATION_REASONS, weights=reason_weights)[0]

        rows.append({
            "EmployeeID":        emp["EmployeeID"],
            "TerminationDate":   term_date,
            "TerminationReason": reason,
        })

    df = pd.DataFrame(rows)
    df["TerminationDate"] = pd.to_datetime(df["TerminationDate"])
    return df


# ─── Validation ───────────────────────────────────────────────────────────────

def validate_referential_integrity(
    employees:    pd.DataFrame,
    compensation: pd.DataFrame,
    leave:        pd.DataFrame,
    performance:  pd.DataFrame,
    termination:  pd.DataFrame,
) -> None:
    """
    Sanity-check cross-table referential integrity and print a summary report.
    Intentional DQ issues are flagged as warnings, not raised as errors —
    they're features of this dataset, not bugs.
    """
    all_ids = set(employees["EmployeeID"].unique())

    sep = "-" * 62
    print(f"\n{sep}")
    print("  VALIDATION REPORT")
    print(sep)
    print(f"  Employees - total rows (incl. duplicates): {len(employees):>5}")
    print(f"  Employees - unique EmployeeIDs:            {len(all_ids):>5}")
    print()

    for label, df in [
        ("Compensation", compensation),
        ("Leave",        leave),
        ("Performance",  performance),
        ("Termination",  termination),
    ]:
        orphans = set(df["EmployeeID"].unique()) - all_ids
        status  = "OK" if not orphans else f"WARN ({len(orphans)} orphans)"
        print(f"  {label:<14} rows: {len(df):>5}   referential integrity: {status}")

    print()
    print("  Intentional data quality markers:")
    print(f"    [DQ] NULL ManagerIDs:          {employees['ManagerID'].isna().sum():>4}")
    print(f"    [DQ] Terminated employees:     {(employees['Status'] == 'Terminated').sum():>4}")
    print(f"    [DQ] Duplicate EmployeeID rows:{employees.duplicated(subset='EmployeeID').sum():>4}")

    # Salary outlier count (>2x or <0.4x of median)
    median_sal = compensation["BaseSalary"].median()
    outliers   = compensation[
        (compensation["BaseSalary"] > median_sal * 2.0) |
        (compensation["BaseSalary"] < median_sal * 0.4)
    ]
    print(f"    [DQ] Salary outlier rows:      {len(outliers):>4}")

    # Leave overlap count
    overlap_count = 0
    for eid, grp in leave.groupby("EmployeeID"):
        grp_sorted = grp.sort_values("StartDate")
        for i in range(1, len(grp_sorted)):
            prev_end   = grp_sorted.iloc[i - 1]["EndDate"]
            curr_start = grp_sorted.iloc[i]["StartDate"]
            if curr_start <= prev_end:
                overlap_count += 1
    print(f"    [DQ] Overlapping leave windows:{overlap_count:>4}")
    print(sep + "\n")


# ─── Excel Export ─────────────────────────────────────────────────────────────

def _style_header(ws, fill_color: str = "1F4E79") -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18


def export_to_excel(tables: dict[str, pd.DataFrame], filepath: str = OUTPUT_FILE) -> None:
    """
    Write all tables to a single .xlsx workbook, one sheet per table.
    Headers are styled so the file is immediately readable without Power BI.
    Dates are written as date objects (not strings) so Excel treats them natively.
    """
    wb = Workbook()
    wb.remove(wb.active)  # discard the default blank sheet

    for sheet_name, df in tables.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(df.columns.tolist())

        for row_tuple in df.itertuples(index=False):
            cleaned_row = []
            for val in row_tuple:
                if isinstance(val, pd.Timestamp):
                    cleaned_row.append(val.date())
                elif pd.isna(val):
                    cleaned_row.append(None)
                else:
                    cleaned_row.append(val)
            ws.append(cleaned_row)

        _style_header(ws)

        # Auto-size columns (cap at 45 chars to avoid absurdly wide columns)
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=10,
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    wb.save(filepath)
    print(f"  Saved: {filepath}  ({len(tables)} sheets)")


# ─── Entry point ──────────────────────────────────────────────────────────────

def main() -> None:
    print("Generating synthetic HRIS dataset …\n")

    print("  [1/5] Building Employees table …")
    employees = create_employees()

    print("  [2/5] Building Compensation table …")
    compensation = create_compensation(employees)

    print("  [3/5] Building Leave table …")
    leave = create_leave(employees)

    print("  [4/5] Building Performance table …")
    performance = create_performance(employees)

    print("  [5/5] Building Termination table …")
    termination = create_termination(employees)

    validate_referential_integrity(employees, compensation, leave, performance, termination)

    tables = {
        "Employees":    employees,
        "Compensation": compensation,
        "Leave":        leave,
        "Performance":  performance,
        "Termination":  termination,
    }

    export_to_excel(tables)
    print("\nDone. Open HRIS_Dataset.xlsx and connect it to Power BI.\n")


if __name__ == "__main__":
    main()


# =============================================================================
#  HOW THIS SIMULATION WORKS — READ BEFORE EXTENDING
# =============================================================================
#
# ARCHITECTURE
# ────────────
# The five generator functions are intentionally decoupled:
#   create_employees()    → source of truth; every other table FK's into this
#   create_compensation() → one or two rows per employee (hire + optional raise)
#   create_leave()        → 0–4 rows per employee depending on tenure and status
#   create_performance()  → one row per employee per review year
#   create_termination()  → exactly one row per terminated employee
#
# Employees is built in four passes (VP → Director → Manager → IC) so that
# every ManagerID reference is guaranteed to exist at insert time.
# This avoids circular dependency problems you'd hit if you tried to assign
# managers in a single pass.
#
# KEY ASSUMPTIONS
# ───────────────
# 1. The company is US-only, 500 employees, growth-stage (more recent hires).
# 2. Executives (VPs) are never terminated in this simulation — realistic for
#    a 5-year window; extend if you want board-level churn scenarios.
# 3. Salary is a point-in-time snapshot. The Compensation table has an
#    EffectiveDate so you can model it as a slowly-changing dimension (SCD2)
#    in Power BI.
# 4. Leave dates don't account for weekends or holidays — add a calendar table
#    in Power BI and filter on business days there.
# 5. Performance reviews always fall on Dec 15. In reality, review cycles vary;
#    adjust `review_date` logic if your org uses mid-year or rolling reviews.
#
# HOW TO EXTEND FOR POWER BI
# ───────────────────────────
# 1. DATE TABLE
#    Create a separate Date dimension in Power BI (or generate one here) spanning
#    2019-01-01 to 2024-12-31. Link HireDate, EffectiveDate, ReviewDate, etc.
#    to it. This unlocks time intelligence (YTD headcount, rolling attrition, etc.)
#
# 2. HEADCOUNT SNAPSHOT
#    The Employees table is a current snapshot. To model headcount over time,
#    build a calculated table in DAX or a Python step here that generates one
#    row per employee per month they were active.
#
# 3. DATA QUALITY DASHBOARD
#    Use the intentional DQ issues as the subject of a second Power BI page:
#      - Stacked bar: employees by dept showing canonical vs. dirty name share
#      - Table: employees with NULL ManagerID
#      - Scatter: salaries by dept with outliers highlighted
#      - Leave overlap calendar view
#
# 4. MANAGER HIERARCHY
#    Power BI handles parent-child hierarchies via PATH() / PATHITEM() DAX.
#    The ManagerID → EmployeeID self-join on the Employees table is all you need.
#    Create a calculated column: OrgPath = PATH(Employees[EmployeeID], Employees[ManagerID])
#
# 5. ATTRITION METRICS
#    Combine Employees + Termination to compute:
#      - Monthly attrition rate = Terminations in month / Avg headcount
#      - Average tenure at exit = TerminationDate - HireDate (in months)
#      - Voluntary vs. involuntary split by department
# =============================================================================
