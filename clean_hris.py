"""
HRIS Data Cleaning Script
==========================
Run AFTER generate_hris.py has created HRIS_Dataset.xlsx.

This script walks through every data quality problem in the raw dataset,
fixes what can be fixed automatically, and flags what a human should review.
It then writes a clean version to HRIS_Dataset_Cleaned.xlsx.

What "cleaning" means here:
  - Remove exact/near-duplicate rows
  - Standardize inconsistent text values (department names, termination reasons)
  - Flag outliers rather than deleting them (a flag column is safer than deletion)
  - Add helper columns that make Power BI reports easier to build
  - Print a before/after report so you can see exactly what changed

Run:
    python clean_hris.py
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────────────────────────────────────
# STEP 0: Load the raw data
# ─────────────────────────────────────────────────────────────────────────────
# pd.read_excel() can read every sheet at once when sheet_name=None.
# It returns a dictionary: {"Employees": DataFrame, "Compensation": DataFrame, ...}

print("=" * 60)
print("  HRIS DATA CLEANING PIPELINE")
print("=" * 60)
print()
print("Loading HRIS_Dataset.xlsx ...")

raw = pd.read_excel("HRIS_Dataset.xlsx", sheet_name=None)

employees    = raw["Employees"].copy()
compensation = raw["Compensation"].copy()
leave        = raw["Leave"].copy()
performance  = raw["Performance"].copy()
termination  = raw["Termination"].copy()

print(f"  Loaded {len(employees)} Employees rows")
print(f"  Loaded {len(compensation)} Compensation rows")
print(f"  Loaded {len(leave)} Leave rows")
print(f"  Loaded {len(performance)} Performance rows")
print(f"  Loaded {len(termination)} Termination rows")
print()

# We will collect a summary of every fix made so you can review it at the end.
cleaning_log: list[str] = []

def log(msg: str) -> None:
    """Print a message and save it to the cleaning log."""
    print(f"  {msg}")
    cleaning_log.append(msg)


# =============================================================================
# TABLE 1: EMPLOYEES
# =============================================================================
print("-" * 60)
print("CLEANING: Employees")
print("-" * 60)

# ── Fix 1: Remove duplicate rows ─────────────────────────────────────────────
#
# WHY this matters: Power BI relationships require unique keys. If EmployeeID
# appears twice, Power BI will refuse to create a relationship between tables,
# or worse, double-count employees in your headcount metric.
#
# HOW we detect it: Two rows with the same EmployeeID are duplicates even if
# the email is spelled differently (our generator made email uppercase on dups).
#
# HOW we fix it: Keep the FIRST occurrence of each EmployeeID. The duplicates
# in this dataset only differed in email case, so keeping the first (lowercase)
# version is safe.

before = len(employees)
employees = employees.drop_duplicates(subset="EmployeeID", keep="first")
removed   = before - len(employees)
log(f"[FIX] Removed {removed} duplicate EmployeeID rows  ({before} -> {len(employees)})")

# ── Fix 2: Standardize email to lowercase ────────────────────────────────────
#
# WHY: The duplicates had UPPER-CASE emails. Even after removing the duplicate
# rows, it's good practice to normalize all emails to lowercase so that future
# lookups and joins are case-insensitive.

employees["Email"] = employees["Email"].str.lower().str.strip()
log("[FIX] Normalized all email addresses to lowercase")

# ── Fix 3: Standardize department names ──────────────────────────────────────
#
# WHY: Power BI groups/filters by the exact text value. "Engineering" and "Eng"
# would appear as TWO separate departments in a slicer — confusing for users
# and wrong for any headcount-by-department calculation.
#
# HOW: We define a mapping from every dirty variant to the canonical name.
# str.strip() removes leading/trailing spaces first.

DEPT_MAP = {
    # Engineering variants
    "Eng":            "Engineering",
    "Engg":           "Engineering",
    "engineering":    "Engineering",
    # Human Resources variants
    "HR":             "Human Resources",
    "H.R.":           "Human Resources",
    "HR Dept":        "Human Resources",
    # Sales variants
    "Sales & BD":     "Sales",
    "SALES":          "Sales",
}

employees["Department"] = employees["Department"].str.strip()   # remove spaces
before_unique = employees["Department"].nunique()

# Apply the map; if a value isn't in the map, leave it unchanged (default=value)
employees["Department"] = employees["Department"].replace(DEPT_MAP)
after_unique = employees["Department"].nunique()

log(f"[FIX] Standardized department names: {before_unique} variants -> {after_unique} unique values")
log(f"      Canonical departments: {sorted(employees['Department'].unique())}")

# ── Fix 4: Flag missing ManagerIDs (do NOT delete these rows) ─────────────────
#
# WHY: Employees without a ManagerID are real people in the org — you don't
# want to delete them. But in Power BI, a blank ManagerID will show as an
# unconnected node in an org hierarchy visual, which is confusing.
#
# BETTER APPROACH: Add a flag column so analysts can filter on it or create
# a dedicated "No Manager Assigned" alert in a dashboard.

employees["ManagerID_Missing"] = employees["ManagerID"].isna()
missing_mgr_count = employees["ManagerID_Missing"].sum()
log(f"[FLAG] {missing_mgr_count} employees have no ManagerID -> new column 'ManagerID_Missing' added")

# ── Fix 5: Ensure consistent date format ─────────────────────────────────────
#
# WHY: When Power BI reads Excel, inconsistent date formats can cause a column
# to be treated as text instead of a date. Forcing datetime here prevents that.

employees["HireDate"] = pd.to_datetime(employees["HireDate"], errors="coerce")
null_dates = employees["HireDate"].isna().sum()
if null_dates > 0:
    log(f"[WARN] {null_dates} employees have unparseable HireDates — set to NaT")
else:
    log("[OK]  All HireDate values parsed successfully")

# ── Fix 6: Validate Status values ────────────────────────────────────────────
#
# WHY: If a rogue value like "terminated" (lowercase) or "Actve" (typo)
# slips in, it would appear as a third category in your Status slicer.
#
# HOW: Check what values exist, then standardize anything that doesn't match.

valid_statuses = {"Active", "Terminated"}
bad_status = employees[~employees["Status"].isin(valid_statuses)]
if len(bad_status) > 0:
    log(f"[WARN] {len(bad_status)} rows have unrecognized Status values: {bad_status['Status'].unique()}")
else:
    log("[OK]  All Status values are 'Active' or 'Terminated'")

print()

# =============================================================================
# TABLE 2: COMPENSATION
# =============================================================================
print("-" * 60)
print("CLEANING: Compensation")
print("-" * 60)

# ── Fix 7: Parse dates ────────────────────────────────────────────────────────
compensation["EffectiveDate"] = pd.to_datetime(compensation["EffectiveDate"], errors="coerce")

# ── Fix 8: Flag salary outliers ───────────────────────────────────────────────
#
# WHY: Salary outliers can badly skew your average compensation charts without
# anyone noticing. The right thing is NOT to delete them (you need HR to
# confirm whether the number is a real senior executive package or a data
# entry error). Instead, flag them so they appear in a "Records Needing Review"
# table in your Power BI dashboard.
#
# HOW: We use the Interquartile Range (IQR) method — the same technique Excel's
# box-plot uses. A salary is suspicious if it falls below Q1 - 1.5*IQR
# or above Q3 + 1.5*IQR.

Q1  = compensation["BaseSalary"].quantile(0.25)
Q3  = compensation["BaseSalary"].quantile(0.75)
IQR = Q3 - Q1

low_bound  = Q1 - 1.5 * IQR
high_bound = Q3 + 1.5 * IQR

compensation["SalaryOutlier"] = (
    (compensation["BaseSalary"] < low_bound) |
    (compensation["BaseSalary"] > high_bound)
)
outlier_count = compensation["SalaryOutlier"].sum()
log(f"[FLAG] {outlier_count} salary rows flagged as outliers (outside ${low_bound:,.0f} - ${high_bound:,.0f})")
log(f"       These need HR review — use 'SalaryOutlier' column to filter in Power BI")

# ── Fix 9: Mark the most recent compensation record per employee ───────────────
#
# WHY: Some employees have two compensation rows (initial hire + merit raise).
# In Power BI, if you just connect Compensation to Employees, you'll get
# double-counting in salary charts. The fix is to add an "IsCurrentSalary"
# flag so you can filter down to one row per employee when needed.
#
# HOW: Sort by EmployeeID and EffectiveDate descending, then mark the first
# row in each group as the current record.

compensation = compensation.sort_values(
    ["EmployeeID", "EffectiveDate"], ascending=[True, False]
)
compensation["IsCurrentSalary"] = ~compensation.duplicated(subset="EmployeeID", keep="first")

current_count = compensation["IsCurrentSalary"].sum()
log(f"[FIX] Added 'IsCurrentSalary' column: {current_count} current records out of {len(compensation)} total")

print()

# =============================================================================
# TABLE 3: LEAVE
# =============================================================================
print("-" * 60)
print("CLEANING: Leave")
print("-" * 60)

# ── Fix 10: Parse dates ───────────────────────────────────────────────────────
leave["StartDate"] = pd.to_datetime(leave["StartDate"], errors="coerce")
leave["EndDate"]   = pd.to_datetime(leave["EndDate"],   errors="coerce")

# ── Fix 11: Add a Duration column ────────────────────────────────────────────
#
# WHY: Power BI can subtract dates in DAX, but having Duration pre-calculated
# in the data model is faster and easier to use in measures. It's also
# immediately readable in a table view.

leave["DurationDays"] = (leave["EndDate"] - leave["StartDate"]).dt.days
log(f"[ADD]  'DurationDays' column calculated for all {len(leave)} leave rows")

# ── Fix 12: Flag impossible date ranges (EndDate before StartDate) ────────────
#
# WHY: A leave record where someone "returned before they left" is logically
# impossible and would produce negative duration in reports.

bad_dates = leave[leave["DurationDays"] < 0]
leave["DateRangeInvalid"] = leave["DurationDays"] < 0
if len(bad_dates) > 0:
    log(f"[FLAG] {len(bad_dates)} leave rows have EndDate before StartDate -> 'DateRangeInvalid' flagged")
else:
    log("[OK]  All leave date ranges are valid (EndDate >= StartDate)")

# ── Fix 13: Detect overlapping leave windows per employee ─────────────────────
#
# WHY: An employee can't be on two separate leave types at the same time
# (in most HR systems). Overlapping windows could mean:
#   - A data entry error (submitted the same leave twice)
#   - A system migration issue (records from two systems merged)
# We flag them for review rather than deleting either record.
#
# HOW: Sort each employee's leaves by StartDate. For each leave after the first,
# check if its StartDate falls before the previous leave's EndDate.

leave_sorted = leave.sort_values(["EmployeeID", "StartDate"])
leave["OverlapWithPriorLeave"] = False

for emp_id, grp in leave_sorted.groupby("EmployeeID"):
    idx_list = grp.index.tolist()
    for i in range(1, len(idx_list)):
        prev_end   = leave.loc[idx_list[i - 1], "EndDate"]
        curr_start = leave.loc[idx_list[i],     "StartDate"]
        if pd.notna(prev_end) and pd.notna(curr_start) and curr_start <= prev_end:
            leave.loc[idx_list[i], "OverlapWithPriorLeave"] = True

overlap_count = leave["OverlapWithPriorLeave"].sum()
log(f"[FLAG] {overlap_count} leave rows overlap with a prior leave for the same employee")
log(f"       Use 'OverlapWithPriorLeave' = TRUE to surface these in a data quality report")

# ── Fix 14: Standardize the Approved column ───────────────────────────────────
#
# WHY: Power BI visuals treat "Y" and "y" as different values. Strip and upper.

leave["Approved"] = leave["Approved"].str.strip().str.upper()
log("[FIX] Standardized 'Approved' column to uppercase (Y/N)")

print()

# =============================================================================
# TABLE 4: PERFORMANCE
# =============================================================================
print("-" * 60)
print("CLEANING: Performance")
print("-" * 60)

performance["ReviewDate"] = pd.to_datetime(performance["ReviewDate"], errors="coerce")

# ── Fix 15: Flag duplicate reviews (same employee, same review year) ──────────
#
# WHY: An employee should only have one performance review per year. A duplicate
# means a data entry error — you don't want Power BI averaging two ratings
# and showing 3.5 when one of those rows was entered by mistake.

performance["ReviewYear"] = performance["ReviewDate"].dt.year
dup_reviews = performance[performance.duplicated(subset=["EmployeeID", "ReviewYear"], keep=False)]
performance["DuplicateReview"] = performance.duplicated(
    subset=["EmployeeID", "ReviewYear"], keep=False
)
dup_count = performance["DuplicateReview"].sum()
if dup_count > 0:
    log(f"[FLAG] {dup_count} performance rows are duplicates (same employee + same year)")
else:
    log("[OK]  No duplicate performance reviews detected")

# ── Fix 16: Convert PromotionEligible from True/False to Yes/No ───────────────
#
# WHY: This is optional, but "Yes"/"No" is much more readable in Power BI
# slicers and table visuals than TRUE/FALSE, especially for non-technical
# stakeholders viewing the dashboard.

performance["PromotionEligible"] = performance["PromotionEligible"].map(
    {True: "Yes", False: "No", "True": "Yes", "False": "No"}
).fillna("No")
log("[FIX] Converted 'PromotionEligible' from True/False to Yes/No")

# ── Fix 17: Add a Rating label column ────────────────────────────────────────
#
# WHY: Numeric ratings (1–5) are useful for calculations, but in a dashboard
# table or tooltip, "Exceeds Expectations" is more meaningful than "5".
# Keep the numeric column AND add a label column — use numeric for math,
# label for display.

RATING_LABELS = {
    1: "1 - Needs Improvement",
    2: "2 - Below Expectations",
    3: "3 - Meets Expectations",
    4: "4 - Exceeds Expectations",
    5: "5 - Outstanding",
}
performance["RatingLabel"] = performance["Rating"].map(RATING_LABELS)
log("[ADD]  'RatingLabel' column added for dashboard display")

print()

# =============================================================================
# TABLE 5: TERMINATION
# =============================================================================
print("-" * 60)
print("CLEANING: Termination")
print("-" * 60)

termination["TerminationDate"] = pd.to_datetime(termination["TerminationDate"], errors="coerce")

# ── Fix 18: Standardize termination reasons ───────────────────────────────────
#
# WHY: NULL, blank, "N/A", and "TBD" all mean "we don't know" — but Power BI
# would show four separate slices in a pie chart. Collapse them all into
# "Unknown" so your voluntary/involuntary breakdown stays meaningful.

REASON_CLEANUP = {
    "":        "Unknown",
    "N/A":     "Unknown",
    "TBD":     "Unknown",
    "Unknown": "Unknown",
}

# Fill NULLs with "Unknown" first, then apply the text replacements
termination["TerminationReason"] = (
    termination["TerminationReason"]
    .fillna("Unknown")
    .str.strip()
    .replace(REASON_CLEANUP)
)
unknown_count = (termination["TerminationReason"] == "Unknown").sum()
log(f"[FIX] Standardized {unknown_count} NULL/blank/N/A termination reasons to 'Unknown'")

# ── Fix 19: Cross-table validation — termination date must be after hire date ─
#
# WHY: This is a cross-table check. Power BI can't do this automatically.
# A termination date before the hire date is physically impossible and
# would make tenure calculations return negative numbers.

# Merge hire dates from the clean employees table
term_check = termination.merge(
    employees[["EmployeeID", "HireDate"]],
    on="EmployeeID",
    how="left",
)
bad_term_dates = term_check[
    term_check["TerminationDate"] <= term_check["HireDate"]
]
termination["TermDateBeforeHireDate"] = termination["EmployeeID"].isin(
    bad_term_dates["EmployeeID"]
)
if len(bad_term_dates) > 0:
    log(f"[FLAG] {len(bad_term_dates)} termination dates are on or before the hire date")
else:
    log("[OK]  All termination dates are after the corresponding hire dates")

# ── Fix 20: Calculate tenure at exit (in months) ─────────────────────────────
#
# WHY: Tenure at exit is one of the most common HR metrics. Pre-calculating
# it here is faster than building the DAX in Power BI and makes the data
# self-explanatory when you share the Excel file with stakeholders.

term_with_hire = termination.merge(
    employees[["EmployeeID", "HireDate"]],
    on="EmployeeID",
    how="left",
)
termination["TenureAtExitMonths"] = (
    (term_with_hire["TerminationDate"] - term_with_hire["HireDate"])
    .dt.days / 30.44
).round(1)
log("[ADD]  'TenureAtExitMonths' column added to Termination table")

print()

# =============================================================================
# FINAL CROSS-TABLE CHECKS
# =============================================================================
print("-" * 60)
print("CROSS-TABLE VALIDATION")
print("-" * 60)

# Check: every EmployeeID in child tables exists in Employees
clean_emp_ids = set(employees["EmployeeID"].unique())
for label, df in [
    ("Compensation", compensation),
    ("Leave",        leave),
    ("Performance",  performance),
    ("Termination",  termination),
]:
    orphans = set(df["EmployeeID"].unique()) - clean_emp_ids
    if orphans:
        log(f"[WARN] {label}: {len(orphans)} EmployeeIDs not found in Employees table -> {orphans}")
    else:
        log(f"[OK]  {label}: all EmployeeIDs exist in Employees table")

print()

# =============================================================================
# CLEANING SUMMARY
# =============================================================================
print("=" * 60)
print("  CLEANING SUMMARY")
print("=" * 60)
fixes   = [line for line in cleaning_log if line.startswith("[FIX]")]
flags   = [line for line in cleaning_log if line.startswith("[FLAG]")]
adds    = [line for line in cleaning_log if line.startswith("[ADD]")]
oks     = [line for line in cleaning_log if line.startswith("[OK]")]
warns   = [line for line in cleaning_log if line.startswith("[WARN]")]

print(f"\n  Fixes applied  (data changed):     {len(fixes)}")
print(f"  Flags added    (review needed):    {len(flags)}")
print(f"  Columns added  (new helpers):      {len(adds)}")
print(f"  Checks passed  (no action needed): {len(oks)}")
print(f"  Warnings       (investigate):      {len(warns)}")
print()
print("  Row counts after cleaning:")
print(f"    Employees:    {len(employees):>5}  (removed {before - len(employees)} duplicates)")
print(f"    Compensation: {len(compensation):>5}")
print(f"    Leave:        {len(leave):>5}")
print(f"    Performance:  {len(performance):>5}")
print(f"    Termination:  {len(termination):>5}")
print()

# =============================================================================
# EXPORT: write the cleaned workbook
# =============================================================================
print("-" * 60)
print("EXPORTING: HRIS_Dataset_Cleaned.xlsx")
print("-" * 60)

def style_header(ws, fill_hex: str = "2E7D32") -> None:
    """Green header for the cleaned file (vs. dark blue on the raw file)."""
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=name)
    ws.append(df.columns.tolist())
    for row_tuple in df.itertuples(index=False):
        cleaned_row = []
        for val in row_tuple:
            if isinstance(val, pd.Timestamp):
                cleaned_row.append(val.date())
            elif pd.isna(val):
                cleaned_row.append(None)
            else:
                cleaned_row.append(val)
        ws.append(cleaned_row)
    style_header(ws)
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=10,
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)


wb = Workbook()
wb.remove(wb.active)

# Drop the internal ReviewYear helper column before exporting
performance_export = performance.drop(columns=["ReviewYear"])

tables_to_export = {
    "Employees":    employees,
    "Compensation": compensation,
    "Leave":        leave,
    "Performance":  performance_export,
    "Termination":  termination,
}

for sheet_name, df in tables_to_export.items():
    write_sheet(wb, sheet_name, df)
    print(f"  Written sheet '{sheet_name}' ({len(df)} rows, {len(df.columns)} columns)")

wb.save("HRIS_Dataset_Cleaned.xlsx")

print()
print("=" * 60)
print("  Done! HRIS_Dataset_Cleaned.xlsx is ready.")
print()
print("  Next steps in Power BI:")
print("  1. Open Power BI Desktop")
print("  2. Home -> Get Data -> Excel Workbook")
print("  3. Select HRIS_Dataset_Cleaned.xlsx")
print("  4. Load all 5 sheets")
print("  5. In Model view, create relationships:")
print("     Compensation[EmployeeID] -> Employees[EmployeeID]")
print("     Leave[EmployeeID]        -> Employees[EmployeeID]")
print("     Performance[EmployeeID]  -> Employees[EmployeeID]")
print("     Termination[EmployeeID]  -> Employees[EmployeeID]")
print()
print("  TIP: In Compensation, filter IsCurrentSalary = TRUE")
print("  before building any salary charts.")
print("=" * 60)
